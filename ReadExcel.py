import openpyxl


DataToRead = openpyxl.load_workbook('Data/DataToRead.xlsx')

DataActive = DataToRead.active

NumbersData = []


for row in range(1, DataActive.max_row):
  for col in DataActive.iter_cols(1, DataActive.max_column):
    NumbersData.append(col[row].value)